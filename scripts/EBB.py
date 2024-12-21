import openpyxl
import cx_Oracle
from datetime import datetime



workbook = openpyxl.load_workbook('DisputeReporting.xlsx')
worksheet = workbook['Sheet1']

## Create a dictionary of column names
ColNames = {}
Current  = 0
maxrows=worksheet.max_row
for COL in worksheet.iter_cols(1, worksheet.max_column):
    ColNames[COL[0].value] = Current
    Current += 1

try:
    db = cx_Oracle.connect('fms_schema1/fms_schema1@192.168.3.57:1521/CHARLQA')
    cursor = db.cursor()
    data=[]
    ## Now we can access by column name
    for row_cells in worksheet.iter_rows(min_row=2,max_row=maxrows):
        NIAcctNo=row_cells[ColNames['New Install Account Number']].value.strip()
        NIWONo=row_cells[ColNames['New Install Work Order Number']].value.strip()
        DisputeDT=str(row_cells[ColNames['Datetime Disputed']].value)
        datetime_object = datetime.strptime(DisputeDT, '%Y-%m-%d %H:%M:%S')
        SubAcctNo=row_cells[ColNames['Sub Account Number']].value.strip()
        UCID=row_cells[ColNames['UC ID']].value.strip()
        if not row_cells[ColNames['Disposition']].value is None:
            Disposition=row_cells[ColNames['Disposition']].value.strip()
        else:
            Disposition=""

        ###Use cursor.executemany
        sqlquery="insert into CHTR_DISPUTE_RPT(NEW_INSTALL_ACCT_NO,NEW_INSTALL_WO_NO,DISPUTE_DT,SUB_ACCT_NO,UCID,DISPOSITION,INSERT_DT) values(:1,:2,:3,:4,:5,:6,:7)"
        data=data+[(NIAcctNo,NIWONo,datetime_object,SubAcctNo,UCID,Disposition,datetime.now())]
        #print(data)
    cursor.executemany(sqlquery,data)
    print(cursor.rowcount, "Rows Inserted")
    db.commit()

except cx_Oracle.DatabaseError as e:
        print ('Exception Occured')
        print(e)
 
finally:
        cursor.close()
        db.close()


