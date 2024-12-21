import openpyxl
import cx_Oracle

workbook = openpyxl.load_workbook('MyTest.xlsx')
worksheet = workbook['MyTest']

## Create a dictionary of column names
ColNames = {}
Current  = 0
maxrows=worksheet.max_row
for COL in worksheet.iter_cols(1, worksheet.max_column):
    ColNames[COL[0].value] = Current
    Current += 1

try:
    db = cx_Oracle.connect('fms_schema1/fms_schema1@192.168.3.57:1521/CHARLQA')
    cursor = db.cursor()

    ## Now you can access by column name
    for row_cells in worksheet.iter_rows(min_row=2,max_row=maxrows):
        ##Test first by priting the data that read from excel is correct##
        #print(row_cells[ColNames['Id']].value)
        #print(row_cells[ColNames['FirstName']].value)
        #print(row_cells[ColNames['LastName']].value)

        ###Use cursor.executemany
        sqlquery="insert into test_table(id,first_name,last_name) values(:1,:2,:3)"
        data=[(row_cells[ColNames['Id']].value, row_cells[ColNames['FirstName']].value, row_cells[ColNames['LastName']].value)]
        print(data)
        cursor.executemany(sqlquery,data)
        print(cursor.rowcount, "Rows Inserted")
    db.commit()

except cx_Oracle.DatabaseError as e:
        print ('Exception Occured')
        print(e)
 
finally:
        cursor.close()
        db.close()


