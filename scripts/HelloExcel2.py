from openpyxl import load_workbook
wb = load_workbook(filename = 'MyTest.xlsx')
MyTest=wb.active
sheet_ranges = wb['MyTest']
print(sheet_ranges['A2'].value)
print(sheet_ranges['B2'].value)
print(sheet_ranges['C2'].value)
print("Number of Rows"+str(MyTest.max_row))
print("Jumber of Columns"+str(MyTest.max_column))

start_col = 0 # 'A' column index
end_col = 2 # 'C' column index
for i in range(1, MyTest.max_row+1):
    row = [cell.value for cell in MyTest[i][start_col:end_col+1]]
    print(row) # list of cell values of this row
